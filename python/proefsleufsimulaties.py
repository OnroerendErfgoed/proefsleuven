# -*- coding: utf-8 -*-

import os
import sys
import random
import math
import arcpy
import numpy
import csv

def header_and_iterator(dataset_name):
# Returns a list of column names and an iterator over the same columns"""
 
    data_description = arcpy.Describe(dataset_name)
    fieldnames = [f.name for f in data_description.fields if f.type not in ["Geometry", "Raster", "Blob"]]
    
    def iterator_for_feature():
        cursor = arcpy.SearchCursor(dataset_name)
        row = cursor.next()
        while row:
            yield [getattr(row, col) for col in fieldnames]
            row = cursor.next()
        del row, cursor
    return fieldnames, iterator_for_feature()

def _encode(x):
    if isinstance(x, unicode):
        return x.encode("utf-8")
    else:
        return str(x)

def _encodeHeader(x):
    return _encode(x.replace(".","_"))
       
def export_to_csv(dataset, output, dialect):

# Output the data to a CSV file

# create the output writer
    
    out_writer = csv.writer(open(output, 'wb'), dialect=dialect)
# return the list of field names and field values
    arcpy.AddMessage("Reading Table...")
    header, rows = header_and_iterator(dataset)
    arcpy.AddMessage("Writing CSV File...")

# write the field names and values to the csv file
    out_writer.writerow(map(_encodeHeader, header))
    
    for row in rows:
        out_writer.writerow(map(_encode, row))
        
def export_to_xls(dataset, output):
    
# Attempt to output to an XLS file. If output file is a .xls file, use xlwt.
# if it is a .xlsx, use openpyxl. If the necessary module is not available,
# the tool fails.
# XLWT can be downloaded from http://pypi.python.org/pypi/xlwt
# OPENPYXL can be downloaded from http://pypi.python.org/pypi/openpyxl

    arcpy.AddMessage("Reading Table...")
    header, rows = header_and_iterator(dataset)
    arcpy.AddMessage("Writing Excel File...")

    def _xls():
        try:
            import xlwt
        except ImportError:
            arcpy.AddError("Import of xlwt module failed.\nThe XLWT module can\
            be downloaded from: http://pypi.python.org/pypi/xlwt")
            return
# Make spreadsheet
        workbook = xlwt.Workbook()
        worksheet = workbook.add_sheet(os.path.split(dataset)[1])

#Set up header row, freeze panes
        header_style = xlwt.easyxf("font: bold on; align: horiz center")
        for index, colheader in enumerate(header):
            worksheet.write(0, index, colheader.replace(".", "_"))
        worksheet.set_panes_frozen(True)
        worksheet.set_horz_split_pos(1)
        worksheet.set_remove_splits(True)

# Write rows
        for rowidx, row in enumerate(rows):
            for colindex, col in enumerate(row):
                worksheet.write(rowidx + 1, colindex, col)

# All done
        workbook.save(output)

    def _xlsx():
        try:
            import openpyxl
        except ImportError:
            arcpy.AddError("Import of module failed.\nThe OPENPYXL module can\
            be downloaded from: http://pypi.python.org/pypi/openpyxl")
            return

# create spreadsheet
        wb = openpyxl.Workbook(optimized_write=True)
        ws = wb.create_sheet()

# insert the header row
        ws.append(header)
# write rows
        for row in rows:
            ws.append(row)

# All done
        wb.save(output)
        
    if output.endswith('.xls'):
        rowcount = int(arcpy.GetCount_management(dataset).getOutput(0))
        if rowcount <= 65535 and len(header) <= 255:
            _xls()
        else:
            arcpy.AddError('Table too large to export to .xls.\
            Select .xlsx output for tables larger than 256 fields x 65535 rows.')

    else:
        _xlsx()

def getValueList (inputTable, field):

# http://joelmccune.com/2012/02/14/create-list-of-unique-table-values-using-python-for-arcgis-analysis/
 
    valueList = [] # array to hold list of values collected
    valueSet = set() # set to hold values to test against to get list
    rows = arcpy.SearchCursor(inputTable) # create search cursor
 
    # iterate through table and collect unique values
    for row in rows:
        value = row.getValue(field)
 
        # add value if not already added and not current year
        if value not in valueSet:
            valueList.append(value)
 
        # add value to valueset for checking against in next iteration
        valueSet.add(value)
 
    # return value list
    valueList.sort()
    return valueList

### optioneel: automatisch overschrijven aanzetten

# arcpy.env.overwriteOutput = True

# workspace wordt door gebruiker geselecteerd

myWorkSpace = arcpy.GetParameterAsText(0)

# input shapefile(opgravingsplan) wordt door gebruiker geselecteerd

InputFeatureClass = arcpy.GetParameterAsText(1)

#####
## deze sectie is optioneel, voor als het gewenst is dat overlappende polygonen in de opgravingsplattegrond worden samengevoegd
## er wordt een file geodatabase aangemaakt voor de opslag van de tijdelijke bestanden

#arcpy.CreateFileGDB_management(myWorkSpace, "tmpGDB.gdb")
#tmpGDB = myWorkSpace+"/tmpGDB.gdb/"

## controle of er overlap van polygonen is; zoja dan wordt dit gecorrigeerd
## N.B. dit betekent dat alle overlappende sporen worden opgesplitst alsof ze in hetzelfde vlak liggen
## als dit niet de bedoeling is, dan dienen de sporenkaarten handmatig te worden aangepast

#Intersect_tmp = tmpGDB+"Intersect_tmp"

#arcpy.Intersect_analysis([InputFeatureClass], Intersect_tmp, "ALL", "", "")

#NIntersects = str(arcpy.GetCount_management(Intersect_tmp).getOutput(0))

#if NIntersects > 0:

#   arcpy.AddWarning("\nEr zijn intersecties gevonden, deze zullen worden gecorrigeerd")

#   Union_tmp = tmpGDB+"Union_tmp"

#   arcpy.Union_analysis([InputFeatureClass], Union_tmp)

#   Dissolve_tmp = tmpGDB+"Dissolve_tmp"

#   arcpy.Dissolve_management(Union_tmp, Dissolve_tmp,["Shape_area"], "", "SINGLE_PART", "DISSOLVE_LINES")

#   NDissolve_shapes = str(arcpy.GetCount_management(Dissolve_tmp).getOutput(0)) 

#   arcpy.AddWarning("\nEr zijn "+NDissolve_shapes+" polygonen gecreëerd\n")
#####

# optioneel worden een analysegebied en een analyseveldnaam geselecteerd

ExtentFeatureClass = arcpy.GetParameterAsText(2)
AnalysisField = arcpy.GetParameterAsText(3)
TranslationChoice = arcpy.GetParameterAsText(4)
FixedMidPointX  = arcpy.GetParameterAsText(5)
FixedMidPointY  = arcpy.GetParameterAsText(6)
RotationChoice = arcpy.GetParameterAsText(7)
FixedRotation = arcpy.GetParameterAsText(8)

# parameters voor uitvoeren simulaties

TrenchLength = float(arcpy.GetParameterAsText(9))
TrenchInterval = float(arcpy.GetParameterAsText(10))
TrenchDistance = float(arcpy.GetParameterAsText(11))
TrenchWidth = float(arcpy.GetParameterAsText(12))
TrenchConfiguration = arcpy.GetParameterAsText(13)
NSimulations = int(arcpy.GetParameterAsText(14))

# exportformaat opgeven (standaard .xls)

ExportFormat = arcpy.GetParameterAsText(15)

TrenchCoverage = 100 * (TrenchLength * TrenchWidth) / (TrenchDistance * TrenchInterval)

arcpy.AddMessage("\nDe gekozen configuratie leidt tot een theoretische dekkingsgraad van "+str(TrenchCoverage)+"%\n")

# bepalen extent van analyse

if ExtentFeatureClass <> "":
    AnalysisExtent = arcpy.Describe(ExtentFeatureClass).extent
    AnalysisArea = 0
    with arcpy.da.SearchCursor(ExtentFeatureClass, "SHAPE@AREA") as cursor:
        for row in cursor:
            AnalysisArea = AnalysisArea + row[0]
else:   
    AnalysisExtent = arcpy.Describe(InputFeatureClass).extent
    AnalysisArea = (AnalysisExtent.XMax -AnalysisExtentXmin) * (AnalysisExtent.YMax -AnalysisExtentYmin)
    
StartXMin = AnalysisExtent.XMin
StartYMin = AnalysisExtent.YMin
EndXMax = AnalysisExtent.XMax
EndYMax = AnalysisExtent.YMax

XExtent = EndXMax - StartXMin
YExtent = EndYMax - StartYMin

# aanmaken locatie voor opslag simulatiebestanden als deze al niet bestaat
# opties stippellijn, hagelslag en vierkantsgrid

if TrenchConfiguration == "Hagelslag":
    arcpy.CreateFolder_management(myWorkSpace, "Hagelslag")
    myWorkSpace = myWorkSpace+"/Hagelslag"
elif TrenchConfiguration == "Stippellijn":
    arcpy.CreateFolder_management(myWorkSpace, "Stippellijn")
    myWorkSpace = myWorkSpace+"/Stippellijn"
elif TrenchConfiguration == "Vierkantsgrid":
    arcpy.CreateFolder_management(myWorkSpace, "Vierkantsgrid")
    myWorkSpace = myWorkSpace+"/Vierkantsgrid"

### N.B. de mapnaam wordt nu in floats opgegeven, omdat het mogelijk moet zijn om de invoer parameters in floats op te geven
### dit evt. nog aanpassen zodat overbodige getallen achter de komma niet worden weergegeven

FolderName = "/L"+str(TrenchLength)+"_I"+str(TrenchInterval)+"_D"+str(TrenchDistance)+"_W"+str(TrenchWidth)+"_"+str(NSimulations)+"Sims"

#if os.path.exists(myWorkSpace+FolderName):
#   exit()

arcpy.CreateFolder_management(myWorkSpace, FolderName)
myWorkSpace = myWorkSpace+FolderName

# er wordt een file geodatabase aangemaakt voor de opslag van de simulatiebestanden

arcpy.CreateFileGDB_management(myWorkSpace, "simGDB.gdb")
simGDB = myWorkSpace+"/simGDB.gdb/"

# er wordt een table gecreërd voor metadata, omvang analysegebied en opslag simulatieresultaten

SimMetaData = simGDB+"SimMetaData"
SimResult = simGDB+"SimResult"
SimExtent = simGDB+"SimExtent"

arcpy.CreateTable_management(simGDB, "SimMetaData", "", "")
arcpy.CreateTable_management(simGDB, "SimResult", "", "")

arcpy.AddField_management(SimMetaData, "ITEM", "TEXT", "", "", "", "", "")
arcpy.AddField_management(SimMetaData, "DESCRIPTION", "TEXT", "", "", "", "", "")

arcpy.AddField_management(SimResult, "RUN", "LONG", "", "", "", "", "")
arcpy.AddField_management(SimResult, "TRENCH_AREA", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "TRENCH_P", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "ROTATION", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "MID_X", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "MID_Y", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "F_INT", "LONG", "", "", "", "", "")
arcpy.AddField_management(SimResult, "A_INT", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "F_INT_P", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "A_INT_P", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "N_INT", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "N_INT_R", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_CONF", "TEXT", "", "", "", "", "") 
arcpy.AddField_management(SimResult, "M_L", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_I", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_D", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_W)", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_C", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_PLAN", "TEXT", "", "", "", "", "" )
arcpy.AddField_management(SimResult, "M_EXTENT", "TEXT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_AREA_AN", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_AREA_EX", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_L_AN", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_W_AN", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_X_MIN", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_Y_MIN", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_X_MAX", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_Y_MAX", "FLOAT", "", "", "", "", "")
arcpy.AddField_management(SimResult, "M_F_AN", "LONG", "", "", "", "", "")

if ExtentFeatureClass <> "":
    AnalysisFeatureClass = simGDB+"ClippedFeatures"
    arcpy.Clip_analysis(InputFeatureClass, ExtentFeatureClass, AnalysisFeatureClass, "")

else:
    AnalysisFeatureClass = InputFeatureClass
    
if AnalysisField <> "":
    
    CategoryList = getValueList (AnalysisFeatureClass, AnalysisField)
    NCategories = len(CategoryList)

    for i in range (0, NCategories):
    
            arcpy.AddField_management(SimResult, "CAT_"+str(i)+"_F_INT", "LONG", "", "", "", "", "")
            arcpy.AddField_management(SimResult, "CAT_"+str(i)+"_A_INT", "FLOAT", "", "", "", "", "")
    #       arcpy.AddField_management(SimResult, "CAT_"+str(i)+"_F_INT_P", "FLOAT", "", "", "", "", "")
    #       arcpy.AddField_management(SimResult, "CAT_"+str(i)+"_A_INT_P", "FLOAT", "", "", "", "", "")
    #       percentages per spoorcategorie niet geimplementeerd, valt te herleiden uit andere gegevens indien gewenst
            
arcpy.CreateFeatureclass_management(simGDB, "SimExtent", "POLYGON", "", "DISABLED", "DISABLED", AnalysisFeatureClass)

# vullen metadatabestand

edit = arcpy.da.Editor(simGDB)
edit.startEditing(False, False)

# let op: de metadata wordt als string weggeschreven, dus voor eventuele nabewerkingen in Excel moet dit worden aangepast

SimMetaData_Cursor = arcpy.da.InsertCursor(SimMetaData, ("ITEM", "DESCRIPTION"))
SimMetaData_Cursor.insertRow(("Configuratie", TrenchConfiguration))
SimMetaData_Cursor.insertRow(("Sleuflengte (L)", str(TrenchLength)))
SimMetaData_Cursor.insertRow(("Sleufinterval (I)", str(TrenchInterval)))
SimMetaData_Cursor.insertRow(("Sleufafstand (D)", str(TrenchDistance)))
SimMetaData_Cursor.insertRow(("Sleufbreedte (W)", str(TrenchWidth)))
SimMetaData_Cursor.insertRow(("Dekkingsgraad (C)", str(TrenchCoverage)+"%"))
SimMetaData_Cursor.insertRow(("Opgravingsplan", InputFeatureClass))
SimMetaData_Cursor.insertRow(("Analysegebied", ExtentFeatureClass))
SimMetaData_Cursor.insertRow(("Oppervlakte analysegebied", str(XExtent * YExtent)))
SimMetaData_Cursor.insertRow(("Oppervlakte opgravingsgebied", str(AnalysisArea)))
SimMetaData_Cursor.insertRow(("Lengte analysegebied", str(max(XExtent, YExtent))))
SimMetaData_Cursor.insertRow(("Breedte analysegebied", str(min(XExtent, YExtent))))
SimMetaData_Cursor.insertRow(("Linkeronderhoek X", str(StartXMin)))
SimMetaData_Cursor.insertRow(("Linkeronderhoek Y", str(StartYMin)))
SimMetaData_Cursor.insertRow(("Rechterbovenhoek X", str(EndXMax)))
SimMetaData_Cursor.insertRow(("Rechterbovenhoek Y", str(EndYMax)))
SimMetaData_Cursor.insertRow(("Aantal sporen in analysegebied", str(arcpy.GetCount_management(AnalysisFeatureClass).getOutput(0))))

FeatureArea = 0

with arcpy.da.SearchCursor(AnalysisFeatureClass, ("OID@", "SHAPE@AREA")) as cursor:
    for row in cursor:
        if row[1] > 0:
            FeatureArea = FeatureArea + row[1]

SimMetaData_Cursor.insertRow(("Oppervlakte sporen in analysegebied", str(FeatureArea)))

# indien een veldnaam wordt opgegeven, dan wordt de analyse apart uitgevoerd over de diverse categorieën
# zoniet, dan wordt de analyse uitgevoerd voor alle sporen samen

if AnalysisField <> "":
    arcpy.AddMessage("De volgende veldnaam is gekozen: "+AnalysisField)
    arcpy.AddMessage("met de volgende "+str(NCategories)+" categorieën: ")
    
    for i in range(0, NCategories):
        arcpy.AddMessage(str(CategoryList[i]))
        SimMetaData_Cursor.insertRow(("Cat_"+str(i), str(CategoryList[i]))) 
    
SimExtent_Cursor = arcpy.da.InsertCursor(SimExtent, ["SHAPE@"])

ExtentGeom = arcpy.Array([arcpy.Point(AnalysisExtent.XMin, AnalysisExtent.YMin),
                    arcpy.Point(AnalysisExtent.XMin, AnalysisExtent.YMax),
                    arcpy.Point(AnalysisExtent.XMax, AnalysisExtent.YMax),
                    arcpy.Point(AnalysisExtent.XMax, AnalysisExtent.YMin),
                    arcpy.Point(AnalysisExtent.XMin, AnalysisExtent.YMin)])
Extent = arcpy.Polygon(ExtentGeom)

SimExtent_Cursor.insertRow([Extent])
        
edit.stopEditing(True)

#### START SIMULATIES VANAF HIER

if RotationChoice == "Regelmatig oplopend":
    RotationDegrees = 0
    RotationIncrease = float(180) / float(NSimulations)
    arcpy.AddMessage("Regelmatig oplopende draaihoek gekozen, de draaihoek wordt met "+str(RotationIncrease)+" graden verhoogd bij elke simulatie")
    
for Simulations in range (1, NSimulations + 1):

    arcpy.AddMessage("Start simulatie "+str(Simulations))

    # maak het proefsleuvenplan aan

    # voor elke simulatie wordt een rotatie en translatie bepaald t.o.v. het middelpunt van de gebiedsomvang
    # translatie is maximaal de sleufafstand in de X-richting en het interval in de Y-richting
    # indien de sleufafstand groter is dan het studiegebied in de X-richting, dan wordt de X-translatie beperkt
    # idem indien het interval groter is dan het studiegebied in Y-richting 
    
    if TranslationChoice == "Willekeurig":
        if TrenchDistance > XExtent:
            TranslateX = random.uniform(-0.5,0.5) * XExtent
        else:
            TranslateX = random.uniform(-0.5,0.5) * TrenchDistance
        if TrenchInterval > YExtent:
            TranslateY = random.uniform(-0.5,0.5) * YExtent
        else:
            TranslateY = random.uniform(-0.5,0.5) * TrenchInterval
    else:
        TranslateX = 0
        TranslateY = 0
    
    if TranslationChoice == "Handmatig invoeren":
        XMidPoint = float(FixedMidPointX)
        YMidPoint = float(FixedMidPointY)
    else:
        XMidPoint = 0.5 * (EndXMax - StartXMin) + StartXMin + TranslateX
        YMidPoint = 0.5 * (EndYMax - StartYMin) + StartYMin + TranslateY

    arcpy.AddMessage("Middelpunt: "+str(XMidPoint)+", "+str(YMidPoint))
    
    # om het sleuvenplan snel te kunnen opstellen, worden de coördinaten van de sleufmiddelpunten in een numpy-array opgeslagen
    # de sleuven moet de gehele gebiedsomvang dekken, ook als de sleuven geroteerd worden om het middelpunt van het proefsleuvenplan

    # stap 1: benodigd aantal sleufmiddelpunten bepalen

    # LET OP: de proefsleufafstand (D) is gedefinieerd als de afstand tussen de middellijnen van de proefsleuven in de dwarsrichting ("kolommen", NCols)
    # Het interval (I) is gedefinieerd als de afstand (in de lengterichting van de sleuven) tussen de proefsleufrijen ("rijen", NRows)
    
    # Bij stippellijn- en hagelslagpatroon verspringen de sleufrijen
    # Daarom is de afstand tussen de kolommen steeds 0.5 * D en de afstand tussen de rijen 2 * I
    
    # Bij het vierkantsgrid is er geen verspringing, en worden gewoon de afstanden D en I aangehouden

    # Het aantal benodigde sleuven wordt afgerond naar boven, en krijgt een veiligheidsmarge van 1 extra sleuf aan alle zijden

    maxExtent = max(XExtent, YExtent)
    maxExtent = math.sqrt(maxExtent * maxExtent + maxExtent * maxExtent)
    
    if TrenchConfiguration == "Stippellijn" or TrenchConfiguration == "Hagelslag":
        NCols = 2 * (int(maxExtent / TrenchDistance + 1) + 2)
        NRows = int(0.5 * maxExtent / (TrenchInterval) + 1) + 2
    else:
        NCols = (int(maxExtent / TrenchDistance + 1) + 2)
        NRows = int(maxExtent / (TrenchInterval) + 1) + 2
    
    # de mogelijke draaihoek ligt tussen 0 en 180 graden; draaihoeken tussen 180 en 360 graden zijn voor rechthoekige objecten gelijk aan die tussen 0 en 180

    if RotationChoice == "Willekeurig":
        RotationDegrees = random.uniform(0,180)
    elif RotationChoice == "Regelmatig oplopend":
        RotationDegrees = RotationDegrees + RotationIncrease
    else:
        arcpy.AddMessage(RotationChoice)
        RotationDegrees = float(FixedRotation)
    
    RotationRadians = math.radians(RotationDegrees)
    Cos = math.cos(RotationRadians)
    Sin = math.sin(RotationRadians)

    arcpy.AddMessage("Draaihoek: "+str(RotationDegrees)+" graden")

    # array aanmaken om de sleufmiddelpunten in op te slaan bij draaihoek 0; deze array is alleen ter controle, en wordt verder niet gebruikt
    # als de sleufrijen verspringen moet hiermee rekening worden gehouden bij het bepalen van de positie van de sleufmiddelpunten

    MidPoints = numpy.zeros((2, NCols * NRows))
    
    if TrenchConfiguration == "Stippellijn" or TrenchConfiguration == "Hagelslag":  
        XPos0 = XMidPoint - (((NCols - 1) * 0.5 * TrenchDistance) / 2)
        YPos0 = YMidPoint - ((NRows - 1) * TrenchInterval) - 0.5 * TrenchInterval
        ColType = "Odd"
    else:
        XPos0 = XMidPoint - (((NCols - 1) * TrenchDistance) / 2)
        YPos0 = YMidPoint - (((NRows - 1) * TrenchInterval) / 2)
        
    XPos = XPos0
    YPos = YPos0
     
    for i in range (0, NCols):
    
        for j in range (0, NRows):
            MidPoints[0, i*NRows + j] = XPos
            MidPoints[1, i*NRows + j] = YPos
            if TrenchConfiguration == "Stippellijn" or TrenchConfiguration == "Hagelslag":  
                YPos = YPos + 2 * TrenchInterval
            else:
                YPos = YPos + TrenchInterval
                
        if TrenchConfiguration == "Stippellijn" or TrenchConfiguration == "Hagelslag":      
            XPos = XPos + 0.5 * TrenchDistance
            
            if ColType == "Odd":
                ColType = "Even"
                YPos = YPos0 + TrenchInterval
            else:
                ColType = "Odd"
                YPos = YPos0        
        else:
            XPos = XPos + TrenchDistance
            YPos = YPos0
            
    # array aanmaken voor de sleufmiddelpunten geroteerd rondom het middelpunt van het proefsleuvenplan
    # N.B. rotatie vindt plaats tegen de klok in
        
    RotatedMidPoints = numpy.zeros((2, NCols * NRows))

    # afstand eerste punt linksonder t.o.v. middelpunt bepalen

    if TrenchConfiguration == "Stippellijn" or TrenchConfiguration == "Hagelslag":
        XOffset = -1 * (((NCols - 1) * 0.5 * TrenchDistance) / 2)
        YOffset0 = -1 * ((NRows - 1) * TrenchInterval) - 0.5 * TrenchInterval
        ColType = "Odd"
    else:
        XOffset = -1 * (((NCols - 1) * TrenchDistance) / 2)
        YOffset0 = -1 * (((NRows - 1) * TrenchInterval) / 2)
        
    YOffset = YOffset0
        
    RotatedXOffset = Cos * XOffset - Sin * YOffset
    RotatedYOffset = Sin * XOffset + Cos * YOffset

    for i in range (0, NCols):
    
        for j in range (0, NRows):
            RotatedMidPoints[0, i*NRows + j] = XMidPoint + RotatedXOffset
            RotatedMidPoints[1, i*NRows + j] = YMidPoint + RotatedYOffset
            
            if TrenchConfiguration == "Stippellijn" or TrenchConfiguration == "Hagelslag":
                YOffset = YOffset + 2 * TrenchInterval
            else:
                YOffset = YOffset + TrenchInterval
            
            RotatedXOffset = Cos * XOffset - Sin * YOffset
            RotatedYOffset = Sin * XOffset + Cos * YOffset

        if TrenchConfiguration == "Stippellijn" or TrenchConfiguration == "Hagelslag":
            XOffset = XOffset + 0.5 * TrenchDistance
            
            if ColType == "Odd":
                ColType = "Even"
                YOffset = YOffset0 + TrenchInterval
                YOffset0 = YOffset
            else:
                ColType = "Odd"
                YOffset = YOffset0 - TrenchInterval
                YOffset0 = YOffset
                
        else:
            XOffset = XOffset + TrenchDistance
            YOffset = YOffset0
            
        RotatedXOffset = Cos * XOffset - Sin * YOffset
        RotatedYOffset = Sin * XOffset + Cos * YOffset

    # aanmaken benodigde basisbestanden voor simulatie
    # SimPoint_0 en SimPoint_1 zijn niet per se nodig, worden alleen ter controle opgeslagen
    
    if Simulations < 10:
        strSimulations = "00" + str(Simulations)
    elif Simulations < 100:
        strSimulations = "0" + str(Simulations)
    else:
        strSimulations = str(Simulations)
        
    SimPoint_0 = simGDB+"Sim"+strSimulations+"_Point_0"
    SimPoint_1 = simGDB+"Sim"+strSimulations+"_Point_1"
    SimPlan_1 = simGDB+"Sim"+strSimulations+"_Plan_1"

    arcpy.CreateFeatureclass_management(simGDB, "Sim"+strSimulations+"_Point_0", "POINT", "", "DISABLED", "DISABLED", AnalysisFeatureClass)
    arcpy.CreateFeatureclass_management(simGDB, "Sim"+strSimulations+"_Point_1", "POINT", "", "DISABLED", "DISABLED", AnalysisFeatureClass)
    arcpy.CreateFeatureclass_management(simGDB, "Sim"+strSimulations+"_Plan_1", "POLYGON", "", "DISABLED", "DISABLED", AnalysisFeatureClass)
        
    arcpy.AddField_management(SimPlan_1, "TrenchNumber", "LONG", "", "", "", "", "")

    # open edit-sessie

    edit = arcpy.da.Editor(simGDB)
    edit.startEditing(False, False)

    # puntbestand met niet-geroteerde sleufmiddelpunten aanmaken

    SimPoint_0_Cursor = arcpy.da.InsertCursor(SimPoint_0, ["SHAPE@XY"])

    for i in range (0, NRows * NCols):
        Point = arcpy.Point(MidPoints[0,i], MidPoints[1,i])
        SimPoint_0_Cursor.insertRow([Point])
        
    # puntbestand met geroteerde sleufmiddelpunten aanmaken

    SimPoint_1_Cursor = arcpy.da.InsertCursor(SimPoint_1, ["SHAPE@XY"])

    for i in range (0, NRows * NCols):

        Point = arcpy.Point(RotatedMidPoints[0,i], RotatedMidPoints[1,i])

        SimPoint_1_Cursor.insertRow([Point])

    # bepaal de positie van de sleufhoekpunten ten opzichte van het sleufmiddelpunt

    LLOffsetX = -0.5 * TrenchWidth
    LLOffsetY = -0.5 * TrenchLength
    ULOffsetX = -0.5 * TrenchWidth
    ULOffsetY = 0.5 * TrenchLength
    UROffsetX = 0.5 * TrenchWidth
    UROffsetY = 0.5 * TrenchLength
    LROffsetX = 0.5 * TrenchWidth
    LROffsetY = -0.5 * TrenchLength

    # roteer de sleufhoekpunten ten opzichte van sleufmiddelpunt
    # N.B. rotatie vindt plaats tegen de klok in
    # voor het hagelslagpatroon moeten de sleuven om en om 90 graden gedraaid worden 

    RotatedLLOffsetX = Cos * LLOffsetX - Sin * LLOffsetY
    RotatedLLOffsetY = Sin * LLOffsetX + Cos * LLOffsetY 
    RotatedULOffsetX = Cos * ULOffsetX - Sin * ULOffsetY
    RotatedULOffsetY = Sin * ULOffsetX + Cos * ULOffsetY
    RotatedUROffsetX = Cos * UROffsetX - Sin * UROffsetY
    RotatedUROffsetY = Sin * UROffsetX + Cos * UROffsetY
    RotatedLROffsetX = Cos * LROffsetX - Sin * LROffsetY
    RotatedLROffsetY = Sin * LROffsetX + Cos * LROffsetY

    if TrenchConfiguration == "Hagelslag":

        PerpendicularRotation = RotationDegrees + 90
        PerpendicularRadians = math.radians(PerpendicularRotation)
        PerpendicularCos = math.cos(PerpendicularRadians)
        PerpendicularSin = math.sin(PerpendicularRadians)
        
        PerpendicularLLOffsetX = PerpendicularCos * LLOffsetX - PerpendicularSin * LLOffsetY
        PerpendicularLLOffsetY = PerpendicularSin * LLOffsetX + PerpendicularCos * LLOffsetY 
        PerpendicularULOffsetX = PerpendicularCos * ULOffsetX - PerpendicularSin * ULOffsetY
        PerpendicularULOffsetY = PerpendicularSin * ULOffsetX + PerpendicularCos * ULOffsetY
        PerpendicularUROffsetX = PerpendicularCos * UROffsetX - PerpendicularSin * UROffsetY
        PerpendicularUROffsetY = PerpendicularSin * UROffsetX + PerpendicularCos * UROffsetY
        PerpendicularLROffsetX = PerpendicularCos * LROffsetX - PerpendicularSin * LROffsetY
        PerpendicularLROffsetY = PerpendicularSin * LROffsetX + PerpendicularCos * LROffsetY
        
    # polygonenbestand met geroteerde sleuven

    SimPlan_1_Cursor = arcpy.da.InsertCursor(SimPlan_1, ["SHAPE@"])

    # in het geval van een hagelslagpatroon moet eerst de startrotatierichting worden vastgesteld

    HagelslagStartRotation = random.randint(0,1)

    for i in range (0, NRows * NCols):

        TrenchMidPointX = RotatedMidPoints[0,i]
        TrenchMidPointY = RotatedMidPoints[1,i]
        
    # plaats de sleufhoekpunten ten opzicht van middelpunt sleuf
    # rotatie hagelslagpatroon moet per sleuf in de rij gewijzigd worden
        
        if TrenchConfiguration == "Hagelslag":
        
            RowChecker = i % (NRows * 2)
            
            if RowChecker > 0:
                if HagelslagStartRotation == 1:
                    HagelslagStartRotation = 0
                else:
                    HagelslagStartRotation = 1
            
            if HagelslagStartRotation == 1:
            
                Point1X = TrenchMidPointX + PerpendicularLLOffsetX
                Point1Y = TrenchMidPointY + PerpendicularLLOffsetY
                Point2X = TrenchMidPointX + PerpendicularULOffsetX
                Point2Y = TrenchMidPointY + PerpendicularULOffsetY
                Point3X = TrenchMidPointX + PerpendicularUROffsetX
                Point3Y = TrenchMidPointY + PerpendicularUROffsetY
                Point4X = TrenchMidPointX + PerpendicularLROffsetX
                Point4Y = TrenchMidPointY + PerpendicularLROffsetY
            
            else: 
                
                Point1X = TrenchMidPointX + RotatedLLOffsetX
                Point1Y = TrenchMidPointY + RotatedLLOffsetY
                Point2X = TrenchMidPointX + RotatedULOffsetX
                Point2Y = TrenchMidPointY + RotatedULOffsetY
                Point3X = TrenchMidPointX + RotatedUROffsetX
                Point3Y = TrenchMidPointY + RotatedUROffsetY
                Point4X = TrenchMidPointX + RotatedLROffsetX
                Point4Y = TrenchMidPointY + RotatedLROffsetY
            
        else:

            Point1X = TrenchMidPointX + RotatedLLOffsetX
            Point1Y = TrenchMidPointY + RotatedLLOffsetY
            Point2X = TrenchMidPointX + RotatedULOffsetX
            Point2Y = TrenchMidPointY + RotatedULOffsetY
            Point3X = TrenchMidPointX + RotatedUROffsetX
            Point3Y = TrenchMidPointY + RotatedUROffsetY
            Point4X = TrenchMidPointX + RotatedLROffsetX
            Point4Y = TrenchMidPointY + RotatedLROffsetY

    # aanmaken sleufpolygoon, en toevoegen aan het sleuvenplan

        TrenchGeom = arcpy.Array([arcpy.Point(Point1X, Point1Y),
                            arcpy.Point(Point2X, Point2Y),
                            arcpy.Point(Point3X, Point3Y),
                            arcpy.Point(Point4X, Point4Y),
                            arcpy.Point(Point1X, Point1Y)])
        Trench = arcpy.Polygon(TrenchGeom)

        SimPlan_1_Cursor.insertRow([Trench])
    
    cursor = arcpy.UpdateCursor(SimPlan_1)
    
    i = 1
    row = cursor.next()
    
    # voeg veld 'TrenchNumber' toe om elke sleuf een individueel nummer te geven
            
    while row:
            
        row.setValue("TrenchNumber", i)
        cursor.updateRow(row)
                    
        i = i + 1
        row = cursor.next()
            
    del row, cursor
    
    # snijd het sleuvenplan af op de omvang van het analysegebied

    SimPlanClip_1 = simGDB+"Sim"+strSimulations+"_PlanClip_1"
    
    if ExtentFeatureClass == "":
        arcpy.Clip_analysis(SimPlan_1, SimExtent, SimPlanClip_1, "")
    else:
        arcpy.Clip_analysis(SimPlan_1, ExtentFeatureClass, SimPlanClip_1, "")   
    
    # bepaal de totale sleufoppervlakte en het sleufdekkingspercentage in het analysegebied
    
    TrenchArea = 0
    
    SC1Fields = ["SHAPE@AREA"]
    
    SC1 = arcpy.da.SearchCursor(SimPlanClip_1, SC1Fields)
    
    for row in SC1:
        
        TrenchArea = TrenchArea + row[0]
    
    TrenchProportion = TrenchArea / AnalysisArea

    # snijd de sporen af binnen het sleuvenplan, en zorg dat de sleufnummers aan de afgesneden sporen worden gekoppeld
    
    SimPlanClip_2 = simGDB+"Sim"+strSimulations+"_PlanClip_2"
    SimPlanClip_3 = simGDB+"Sim"+strSimulations+"_PlanClip_3"

    arcpy.Clip_analysis(AnalysisFeatureClass, SimPlanClip_1, SimPlanClip_2, "")
    arcpy.SpatialJoin_analysis(SimPlanClip_2, SimPlanClip_1, SimPlanClip_3)
    
    # tel aantal doorsneden sporen en doorsneden oppervlakte

    ClippedFeatures = int(arcpy.GetCount_management(SimPlanClip_3).getOutput(0))
    ClippedFeaturesIntersectionCount = 0
    with arcpy.da.SearchCursor(SimPlanClip_3, "Join_Count") as cursor:     
        for row in cursor:                                                      
            ClippedFeaturesIntersectionCount = ClippedFeaturesIntersectionCount + row[0] 
    TotalFeatures = int(arcpy.GetCount_management(AnalysisFeatureClass).getOutput(0))
    ClippedFeaturesProportion = float(ClippedFeatures) / float(TotalFeatures)
    ClippedFeaturesPercentage = float(100 * ClippedFeaturesProportion)
    ClippedFeaturesIntersectionCountRatio =  float(ClippedFeaturesIntersectionCount) / float(TotalFeatures)

    arcpy.AddMessage("Aantal doorsneden sporen: "+str(ClippedFeatures)+" van "+str(TotalFeatures))
    
    # let op hier, de arcpy.da.SearchCursor functie werkte niet naar behoren, want de cursor kon niet worden teruggezet op 0
    # voor hergebruik (mogelijk een bug in ArcGIS 10.1?)

    # daarom worden de waarden van de spoorcategorieën en bijbehorende oppervlakten in een Python-array opgeslagen
    # als AnalysisField niet is ingevuld, dan wordt deze stap overgeslagen

    ClippedArea = 0
    
    if AnalysisField <> "":
    
        SC2Fields = ["SHAPE@AREA", AnalysisField]
        CategoryValues = [[],[]]
    
    else:
        
        SC2Fields = ["SHAPE@AREA"]
    
    SC2 = arcpy.da.SearchCursor(SimPlanClip_3, SC2Fields)

    for row in SC2:
    
        if AnalysisField <> "":
            
            CategoryValues[0].append(row[0])
            CategoryValues[1].append(row[1])
        
        ClippedArea = ClippedArea + row[0]
    
    ClippedAreaProportion = ClippedArea / FeatureArea
    ClippedAreaPercentage = float(100 * ClippedAreaProportion)
    arcpy.AddMessage("Doorsneden oppervlakte: "+str(ClippedArea)+" van "+str(FeatureArea)+", ofwel "+str(ClippedAreaPercentage)+"%")
    
    # voeg de resultaten van de simulatie toe aan SimResult

    SimResult_Cursor = arcpy.da.InsertCursor(SimResult, 
                                                        (
                                                         "RUN", 
                                                         "TRENCH_AREA", 
                                                         "TRENCH_P", 
                                                         "ROTATION", 
                                                         "MID_X", 
                                                         "MID_Y", 
                                                         "F_INT", 
                                                         "A_INT", 
                                                         "F_INT_P", 
                                                         "A_INT_P", 
                                                         "N_INT", 
                                                         "N_INT_R",
                                                         "M_CONF",
                                                         "M_L",
                                                         "M_I",
                                                         "M_D",
                                                         "M_W",
                                                         "M_C",
                                                         "M_PLAN",
                                                         "M_EXTENT",
                                                         "M_AREA_AN",
                                                         "M_AREA_EX",
                                                         "M_L_AN", 
                                                         "M_W_AN",
                                                         "M_X_MIN",
                                                         "M_Y_MIN",
                                                         "M_X_MAX",
                                                         "M_Y_MAX",
                                                         "M_F_AN"
                                                        )
                                             )
    SimResult_Cursor.insertRow((Simulations,
                                    TrenchArea,
                                    TrenchProportion,
                                    RotationDegrees,
                                    XMidPoint,
                                    YMidPoint,
                                    ClippedFeatures,
                                    ClippedArea,
                                    ClippedFeaturesPercentage,
                                    ClippedAreaPercentage,
                                    ClippedFeaturesIntersectionCount,
                                    ClippedFeaturesIntersectionCountRatio,
                                    TrenchConfiguration,
                                    TrenchLength,
                                    TrenchInterval,
                                    TrenchDistance,
                                    TrenchWidth,
                                    TrenchCoverage,
                                    InputFeatureClass,
                                    ExtentFeatureClass,
                                    XExtent * YExtent,
                                    AnalysisArea,
                                    max(XExtent, YExtent),
                                    min(XExtent, YExtent),
                                    StartXMin,
                                    StartYMin,
                                    EndXMax,
                                    EndYMax,
                                    arcpy.GetCount_management(AnalysisFeatureClass).getOutput(0)
                                 ))
                                    
    edit.stopEditing(True)
                                    
    # indien een analyseveld is gekozen, voeg de resultaten van de analyse per spoorcategorie toe
                                    
    if AnalysisField <> "":
        
        for i in range(0, NCategories):
            
            CatClippedFeatures = 0
            CatClippedArea = 0
            arcpy.AddMessage(str(CategoryList[i]))
 
            for j in range(0, ClippedFeatures):
                
                PolyCategory = CategoryValues[1][j]
                PolyArea = CategoryValues[0][j]
                
                if PolyCategory == str(CategoryList[i]):
                    CatClippedFeatures = CatClippedFeatures + 1
                    CatClippedArea = CatClippedArea + PolyArea
                    
            arcpy.AddMessage("Sporen: "+str(CatClippedFeatures))
            arcpy.AddMessage("Oppervlakte: "+str(CatClippedArea))
                
            fc = SimResult
            queryString = '"' + "RUN" + '" = ' + str(Simulations)
            field1 = "RUN"
            field2 = "CAT_"+str(i)+"_F_INT"
            field3 = "CAT_"+str(i)+"_A_INT"

            cursor = arcpy.UpdateCursor(fc, queryString)
            
            row = cursor.next()
            
            while row:
            
                row.setValue(field2, CatClippedFeatures)
                row.setValue(field3, CatClippedArea)
                cursor.updateRow(row)
                    
                row = cursor.next()
            
            del row, cursor
    
    arcpy.Delete_management("in_memory")

edit.startEditing(False, False)

### afsluiten simulaties

SimMetaData_Cursor = arcpy.da.InsertCursor(SimMetaData, ("ITEM", "DESCRIPTION"))
SimMetaData_Cursor.insertRow(("Aantal runs", str(Simulations)))

edit.stopEditing(True)

# exporteren resultaten

if ExportFormat == ".csv":

    dataset_name1 = SimResult
    dataset_name2 = SimMetaData 
    output_file1 = myWorkSpace+"\SimResult.csv"
    output_file2 = myWorkSpace+"\SimMetaData.csv"
    dialect = 'excel'
    try:
        export_to_csv(dataset_name1, output_file1, dialect)
        export_to_csv(dataset_name2, output_file2, dialect)
    except Exception as err:
        arcpy.AddError('Error: {0}'.format(err))
        
elif ExportFormat == ".xls":
    dataset_name1 = SimResult
    dataset_name2 = SimMetaData 
    output_file1 = myWorkSpace+"\SimResult.xls"
    output_file2 = myWorkSpace+"\SimMetaData.xls"
    try:
        export_to_xls(dataset_name1, output_file1)
        export_to_xls(dataset_name2, output_file2)
    except Exception as err:
        arcpy.AddError("Error: {0}".format(err))
        
elif ExportFormat == ".xlsx":
    dataset_name1 = SimResult
    dataset_name2 = SimMetaData 
    output_file1 = myWorkSpace+"\SimResult.xlsx"
    output_file2 = myWorkSpace+"\SimMetaData.xlsx"
    try:
        export_to_xls(dataset_name1, output_file1)
        export_to_xls(dataset_name2, output_file2)
    except Exception as err:
        arcpy.AddError("Error: {0}".format(err))
